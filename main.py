#!/usr/bin/env python3
"""This is the main module"""
import requests
import openpyxl
from dotenv import load_dotenv
import os

load_dotenv()

API_KEY = os.getenv("API_KEY")  # Replace with your actual API key
CSE_ID = os.getenv("CSE_ID")    # Replace with your actual Custom Search Engine ID

def google_search(query, api_key, cse_id, start_index=1, num_results=10):
    url = "https://www.googleapis.com/customsearch/v1"
    params = {
        'q': query,
        'key': api_key,
        'cx': cse_id,
        'start': start_index,
        'num': num_results
    }
    response = requests.get(url, params=params)
    return response.json()

def extract_domains_from_results(results):
    domains = set()
    for item in results.get('items', []):
        domain = item['link'].split("//")[0] + "//" + item['link'].split("//")[-1].split("/")[0]
        domains.add(domain)
    return list(domains)

def save_to_excel(domains, filename="search_results.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Domains"
    ws.cell(row=1, column=1, value="Domain")
    for idx, domain in enumerate(domains, 2):
        ws.cell(row=idx, column=1, value=domain)
    wb.save(filename)

def main():
    keyword = input("Enter the keyword for the search: ")
    all_domains = set()
    for page in range(1, 101, 10):  # To get results from the first 10 pages
        results = google_search(keyword, API_KEY, CSE_ID, start_index=page)
        domains = extract_domains_from_results(results)
        all_domains.update(domains)
    save_to_excel(list(all_domains))
    print(f"Saved {len(all_domains)} domains to 'search_results.xlsx'.")

if __name__ == "__main__":
    main()
